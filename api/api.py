from flask import Flask, send_file, request, abort
import configparser
import os
import re
import sys
import openpyxl
import requests
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from requests import JSONDecodeError

app = Flask(__name__)

@app.route('/', methods=['GET'])
def index():
	return send_file('index.html')

@app.route('/api/get_furniture_list', methods=['GET'])
def api_get_furniture_list():
	uid = request.args.get('uid')
	cookie = request.args.get('cookie')
	share_code = request.args.get('share_code')
	
	if not all([uid, cookie, share_code]):
		abort(400, 'Missing required parameters: uid, cookie, or share_code')
	
	json_data = get_json(share_code, cookie)
	res = parse_json(json_data)
	
	filename = f'{uid}_{share_code}.xlsx'
	out_excel(res, filename)
	
	try:
		return send_file(filename, as_attachment=True)
	except FileNotFoundError:
		abort(404, 'File not found.')


def get_json(share_code, cookie):
	url = 'https://api-takumi.mihoyo.com/event/e20200928calculate/v1/furniture/blueprint'
	params = {
		'share_code': share_code,
		'region': 'cn_gf01'
	}
	headers = {
		'cookie': cookie,
		'Referer': 'https://webstatic.mihoyo.com/'
	}
	proxies = {"http": None, "https": None}
	resp = requests.get(url, params, headers=headers, proxies=proxies)
	try:
		return resp.json()
	except JSONDecodeError:
		print(resp.content)
		sys.exit('解析出错，接口返回不符合预期')


def parse_json(data):
	if data['retcode'] == -100:
		sys.exit('cookie已过期，请重新登录后获取')
	elif not data['retcode'] == 0:
		sys.exit(data['message'])
	res = []
	for furniture in data['data']['list'] + data['data']['not_calc_list']:
		jump = ''
		if furniture["wiki_url"] != "":
			url = re.sub(r'\?.*$', '', furniture["wiki_url"])
			jump = '=HYPERLINK(\"' + url + '\", \"跳转\")'
		res.append([furniture['id'], furniture['name'], "", furniture['num'], furniture['level'], jump])
	res.sort(key=lambda cur_res: (cur_res[4], cur_res[3], cur_res[0]))
	return res


def out_excel(data, filename):
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.append(['摆设ID', '摆设名称', '拥有数量', '所需数量', '星级', 'Wiki'])
	for row in data:
		ws.append(row)
	
	rule = ColorScaleRule(start_type='num', start_value=2, start_color='ffc9e8a8',
	                      mid_type='num', mid_value=3, mid_color='ff6ddbff',
	                      end_type='num', end_value=4, end_color='ffcaa8e5')
	ws.conditional_formatting.add(f'E2:E{len(data) + 1}', rule)
	green_fill = PatternFill(start_color='e2efda',
	                         end_color='e2efda',
	                         fill_type='solid')
	ws.conditional_formatting.add(f'C2:C{len(data) + 1}',
	                              CellIsRule(operator='greaterThanOrEqual', formula=['$D2'], fill=green_fill))
	
	rule = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=max(row[3] for row in data),
	                   color="FFfcd5b4")
	ws.conditional_formatting.add(f'D2:D{len(data) + 1}', rule)
	
	side = Side(style='thin', color='00000000')
	border = Border(*(4 * [side]))
	font = Font(name='等线')
	alignment = Alignment(horizontal='center', vertical='center')
	rows = ws[f'A1:F{len(data) + 1}']
	for row in rows:
		for cell in row:
			cell.border = border
			cell.font = font
			cell.alignment = alignment
	
	ws.column_dimensions['B'].width = 25
	ws.freeze_panes = 'A2'
	
	try:
		wb.save(filename)
	except PermissionError:
		print('请关闭当前Excel表后重试')
		sys.exit()


if __name__ == '__main__':
	app.run(debug=True)
